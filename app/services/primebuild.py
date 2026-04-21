"""Primebuild business logic (Journals, Hours Worked, Keypay Location).

Lifted from Handover_automations/Primebuild/app.py per the 'wrap, don't rewrite'
rule. Only pure functions are copied; Streamlit calls are dropped. File inputs
are adapted to (bytes, filename) tuples so routers can pass UploadFile data.
"""

import io
import re
from datetime import datetime, timedelta

import numpy as np
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


# ═══════════════════════════════════════════════════════════════════════════════
# PAYROLL JOURNALS
# ═══════════════════════════════════════════════════════════════════════════════

STATE_CWI = {'NSW': 10, 'QLD': 40, 'VIC': 20, 'ROL': 11, 'SVS': 85, 'CON': 50}

LOOKUP_K = {
    'CBA Cheque Account', 'PAYG Withholding Tax', 'Payroll Tax Payable',
    'Superannuation Clearing', 'Payroll Clearing', 'Other Payroll Deductions',
    'Annual Leave Entitlements', 'Sick & Personal Leave Entitlem',
    'RDO Accrual', 'LSL Provision', 'Operations SLPHUni Accrual',
    'Provision - WCompensation'
}

OUTPUT_HEADERS = [
    'Dissection', 'Description', 'Costing Work Id', 'Job', 'Cost Code',
    'Cost Type', 'Resource Code', 'GL Account', 'Quantity', 'UOM', 'Amount',
    'Normal Value', 'Allowance Value', 'On Cost Value', 'Tax Code',
    'Tax Percentage', 'Tax Amount', 'Internal Reference', 'External Reference',
    'Asset Work Id', 'Asset', 'Small Order Revenue', 'Text',
]


def parse_filename(filename):
    name = filename.replace('.xlsx', '').replace('.xlsm', '')
    state = name[:3].upper()
    is_wcomp = 'WCOMP' in name.upper()
    freq = 'WC' if is_wcomp else name[4:6].upper()
    cwi = STATE_CWI.get(state, 10)
    return state, freq, cwi


def process_raw_file(file_bytes, filename, payment_date_str):
    state, freq, default_cwi = parse_filename(filename)
    dt = datetime.strptime(payment_date_str, '%d/%m/%Y')
    fmt_date = dt.strftime('%d/%m/%Y')
    time_date_fname = dt.strftime('%Y%m%d')
    upper_name = filename.upper().replace('.XLSX', '')
    if 'WCOMP' in upper_name:
        m = re.search(r'WCOMP_([A-Z]+)_', upper_name)
        raw_freq = m.group(1) if m else 'FN'
    else:
        raw_freq = upper_name[4:6]
    internal_ref = f"{state} {raw_freq} PAY {fmt_date}"
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name='Journal', header=None)
    data = df.iloc[1:].reset_index(drop=True)
    rows = []
    dissection = 0
    for _, row in data.iterrows():
        acct_no   = str(int(row[3])) if pd.notna(row[3]) else ''
        acct_name = str(row[4]) if pd.notna(row[4]) else ''
        desc      = str(row[5]) if pd.notna(row[5]) else ''
        amount    = row[6] if pd.notna(row[6]) else 0
        tax_code  = str(row[10]) if pd.notna(row[10]) and str(row[10]) != 'nan' else ''
        keypay    = str(row[11]) if pd.notna(row[11]) and str(row[11]) != 'nan' else ''
        dim2      = str(row[13]) if pd.notna(row[13]) and str(row[13]) != 'nan' else ''
        dissection += 1
        col_b = f"{desc} {acct_name}".strip()
        in_lookup = acct_name in LOOKUP_K
        n_char4 = dim2[3] if len(dim2) > 3 else ''
        col_c = default_cwi if (in_lookup or n_char4 not in ('C', 'D')) else int(dim2[:2])
        col_d = col_e = col_f = ''
        if len(dim2) >= 10 and '/' in dim2:
            parts     = dim2.split('/')
            cw_id     = parts[0] if len(parts) > 0 else ''
            job_code  = parts[1] if len(parts) > 1 else ''
            cost_code = parts[2] if len(parts) > 2 else ''
            cost_type = parts[3] if len(parts) > 3 else ''
            if cost_type == 'RV':
                cost_type = 'RC'
            if cw_id.isdigit():
                col_c = int(cw_id)
            col_d, col_e, col_f = job_code, cost_code, cost_type
            if state == 'ROL':
                if cost_code.startswith('98'):
                    col_f = cost_type
                elif len(cost_code) > 3 and cost_code[3] == '-':
                    col_f = 'RC'
                elif len(cost_code) > 7:
                    col_f = cost_type
                elif len(cost_code) == 5:
                    col_f = 'LB'
                else:
                    col_f = 'RC'
                if job_code and job_code[0] == 'D' and not cost_code.startswith('98'):
                    col_f = 'LB'
            if job_code and job_code[0] == 'R' and state != 'ROL':
                col_f = 'CA'
        has_job = len(dim2) >= 10 and '/' in dim2
        col_h = '' if has_job else f"{default_cwi}{acct_no}"
        if keypay.lower().startswith('prime build') and 'HOLDINGS' not in keypay:
            right6 = keypay[-6:]
            col_h = right6
            col_c = int(right6[:2]) if right6[:2].isdigit() else col_c
            col_d = col_e = col_f = ''
        if state == 'NSW' and freq == 'WK':
            if 'rollouts' in col_b.lower() or '- ROL' in col_b:
                col_c = 11
        rows.append({
            'Dissection': dissection, 'Description': col_b, 'Costing Work Id': col_c,
            'Job': col_d if col_d else None, 'Cost Code': col_e,
            'Cost Type': col_f if col_f else None, 'Resource Code': '',
            'GL Account': col_h, 'Quantity': None, 'UOM': None, 'Amount': amount,
            'Normal Value': None, 'Allowance Value': None, 'On Cost Value': None,
            'Tax Code': tax_code if tax_code else None, 'Tax Percentage': None,
            'Tax Amount': None, 'Internal Reference': internal_ref,
            'External Reference': str(int(float(row[1]))) if pd.notna(row[1]) else '',
            'Asset Work Id': None, 'Asset': None, 'Small Order Revenue': None,
            'Text': col_b,
        })
    out_name = f"WComp {state} {time_date_fname} JNL.xlsx" if freq == 'WC' \
               else f"{state} {freq} {time_date_fname} JNL.xlsx"
    return rows, out_name, state, freq, default_cwi


def build_journal_workbook(rows, state, freq, cwi, payment_date_str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "JRNDOWNLD"
    dt = datetime.strptime(payment_date_str, '%d/%m/%Y')
    title_font = Font(name='Arial', bold=True, size=20)
    batch_font = Font(name='Arial', size=10)
    hdr_font   = Font(name='Arial', bold=True, size=11, color='FFFFFFFF')
    navy_fill  = PatternFill('solid', fgColor='002856')
    data_font  = Font(name='Arial', size=11)
    row_fill   = PatternFill('solid', fgColor='E5B8B7')
    ws['A1'] = 'General Ledger Journal Download'
    ws['A1'].font = title_font
    ws.merge_cells('A1:Y1')
    for col_i, lbl in enumerate(['Batch', 'Work Id', 'Date', 'Period', 'Batch Type',
                                  'Debit Check Sum', 'Line Count Check', 'Text'], 1):
        ws.cell(row=2, column=col_i, value=lbl).font = batch_font
    ws.cell(row=3, column=2, value=cwi).font = batch_font
    date_cell = ws.cell(row=3, column=3, value=dt)
    date_cell.number_format = 'DD/MM/YYYY'
    date_cell.font = batch_font
    ws.cell(row=3, column=5, value='G').font = batch_font
    col_headers = OUTPUT_HEADERS + ['Error Message', 'POSTED']
    for col_i, hdr in enumerate(col_headers, 1):
        cell = ws.cell(row=4, column=col_i, value=hdr)
        cell.font = hdr_font
        cell.fill = navy_fill
    ws.cell(row=5, column=25, value='POSTED').font = data_font
    total_cols = len(col_headers)
    for row_data in rows:
        r = ws.max_row + 1
        for col_i in range(1, total_cols + 1):
            ws.cell(row=r, column=col_i).fill = row_fill
        for col_i, key in enumerate(OUTPUT_HEADERS, 1):
            val = row_data.get(key)
            cell = ws.cell(row=r, column=col_i, value=val)
            cell.font = data_font
            cell.fill = row_fill
            if key in ('GL Account', 'Cost Code', 'External Reference'):
                cell.number_format = '@'
            if key == 'Amount' and val is not None:
                cell.number_format = '#,##0.00'
    col_widths = {
        'A': 10, 'B': 50, 'C': 16, 'D': 12, 'E': 14, 'F': 10,
        'G': 14, 'H': 14, 'I': 10, 'J': 8,  'K': 14, 'L': 14,
        'M': 16, 'N': 14, 'O': 10, 'P': 14, 'Q': 12, 'R': 24,
        'S': 18, 'T': 14, 'U': 10, 'V': 18, 'W': 50,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = 'A6'
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# HOURS WORKED
# ═══════════════════════════════════════════════════════════════════════════════

LONG_SHIFT_THRESHOLD   = 14.0
SHORT_BREAK_THRESHOLD  = 10.0
WEEKLY_HOURS_THRESHOLD = 60.0
FATIGUE_SUM_THRESHOLD  = 14.0
HW_YELLOW = PatternFill("solid", fgColor="FFFF00")
HW_RED    = PatternFill("solid", fgColor="FF0000")

HEADERS_29 = [
    "Employee Id", "First Name", "Surname", "Employee External Id",
    "Timesheet Id", "Status", "Location",
    "Start Date", "Start Time", "End Date", "End Time",
    None,
    "Actual Start Date", "Actual Start Time", "Actual End Date", "Actual End Time",
    "Time Variance", "Duration", "Total Duration", "Units", "Unit Type",
    "Work Type", "Shift Conditions", "Classification",
    "Number Of Breaks", "Break Duration",
    "Consolidated With Timesheet Line Id", "Reviewed By", "Created Date UTC",
]


def parse_duration(val) -> float:
    if pd.isna(val):
        return 0.0
    s = str(val).strip()
    m = re.match(r'^(\d+):(\d+):(\d+)$', s)
    if m:
        return int(m.group(1)) + int(m.group(2)) / 60 + int(m.group(3)) / 3600
    try:
        return float(s)
    except Exception:
        return 0.0


def process_hours_file(file_bytes, filename) -> dict:
    xl   = pd.ExcelFile(io.BytesIO(file_bytes))
    df   = pd.read_excel(xl, sheet_name="Export", header=0)
    stem = re.sub(r'\.xlsx?$', '', filename, flags=re.I)
    wt   = df["Work Type"].astype(str).str.strip()
    df_shift = df[
        ((wt.str.lower() == "shift work") | (df["Work Type"].isna()) | (wt == "")) &
        (df["Employee Id"].astype(str).str.strip().str.lower() != "totals")
    ].copy()
    if df_shift.empty:
        return {"raw_df": df, "long_shift_df": pd.DataFrame(), "weekly_df": pd.DataFrame(),
                "summary": {}, "filename_stem": stem, "error": "No shift data rows found."}
    def make_dt(date_col, time_col):
        dates = pd.to_datetime(df_shift[date_col], errors="coerce")
        times = df_shift[time_col].apply(
            lambda v: v if isinstance(v, pd.Timedelta) else pd.to_timedelta(str(v), errors="coerce")
        )
        return dates + times
    df_shift["shift_start"]  = make_dt("Start Date", "Start Time")
    df_shift["shift_end"]    = make_dt("End Date",   "End Time")
    df_shift["duration_hrs"] = df_shift["Duration"].apply(parse_duration)
    df_shift["full_name"]    = (df_shift["First Name"].astype(str).str.strip() + " " +
                                df_shift["Surname"].astype(str).str.strip())
    df_shift.sort_values(["full_name", "shift_start"], inplace=True)
    df_shift.reset_index(drop=True, inplace=True)
    df_shift["long_shift_flag"]  = df_shift["duration_hrs"] > LONG_SHIFT_THRESHOLD
    df_shift["break_before_hrs"] = np.nan
    df_shift["short_break_flag"] = False
    for i in range(1, len(df_shift)):
        prev, curr = df_shift.iloc[i - 1], df_shift.iloc[i]
        if prev["full_name"] == curr["full_name"]:
            gap = (curr["shift_start"] - prev["shift_end"]).total_seconds() / 3600
            df_shift.at[df_shift.index[i], "break_before_hrs"] = round(gap, 2)
            if gap < SHORT_BREAK_THRESHOLD:
                df_shift.at[df_shift.index[i], "short_break_flag"] = True
    df_shift["fatigue_risk_flag"] = False
    for i in range(1, len(df_shift)):
        if not df_shift.at[df_shift.index[i], "short_break_flag"]:
            continue
        prev, curr = df_shift.iloc[i - 1], df_shift.iloc[i]
        if prev["full_name"] != curr["full_name"]:
            continue
        combined = (df_shift.at[df_shift.index[i - 1], "duration_hrs"] +
                    df_shift.at[df_shift.index[i], "duration_hrs"])
        if combined > FATIGUE_SUM_THRESHOLD:
            df_shift.at[df_shift.index[i - 1], "fatigue_risk_flag"] = True
            df_shift.at[df_shift.index[i],     "fatigue_risk_flag"] = True
    long_shift_df = (df_shift[df_shift["long_shift_flag"] | df_shift["fatigue_risk_flag"]]
                     .copy().sort_values(["Employee Id", "shift_start"]).reset_index(drop=True))
    weekly = (df_shift.groupby("Employee Id")["duration_hrs"].sum().reset_index()
              .rename(columns={"Employee Id": "Employee", "duration_hrs": "Total Hours"}))
    weekly["Total Hours"] = weekly["Total Hours"].round(2)
    weekly["Exceeds 60h"] = weekly["Total Hours"] > WEEKLY_HOURS_THRESHOLD
    weekly.sort_values("Total Hours", ascending=False, inplace=True)
    summary = {
        "total_employees": df_shift["full_name"].nunique(),
        "total_shifts":    len(df_shift),
        "long_shifts":     int(df_shift["long_shift_flag"].sum()),
        "short_breaks":    int(df_shift["fatigue_risk_flag"].sum()),
        "fatigue_flags":   int(df_shift["fatigue_risk_flag"].sum()),
        "exceed_60h":      int(weekly["Exceeds 60h"].sum()),
    }
    return {"raw_df": df, "shift_df": df_shift, "long_shift_df": long_shift_df,
            "weekly_df": weekly, "summary": summary, "filename_stem": stem}


def _raw_row_values(row: pd.Series) -> list:
    def safe_date(v):
        try:
            return pd.to_datetime(v).date() if pd.notna(v) else None
        except Exception:
            return None
    def safe_td(v):
        if isinstance(v, pd.Timedelta):
            return timedelta(seconds=int(v.total_seconds()))
        if isinstance(v, timedelta):
            return v
        try:
            return timedelta(seconds=int(pd.to_timedelta(str(v)).total_seconds()))
        except Exception:
            return None
    return [
        row.get("Employee Id"), row.get("First Name"), row.get("Surname"),
        row.get("Employee External Id"), row.get("Timesheet Id"), row.get("Status"),
        row.get("Location"),
        safe_date(row.get("Start Date")), safe_td(row.get("Start Time")),
        safe_date(row.get("End Date")),   safe_td(row.get("End Time")),
        "Less than 10 hour break" if row.get("short_break_flag") else "",
        safe_date(row.get("Actual Start Date")), safe_td(row.get("Actual Start Time")),
        safe_date(row.get("Actual End Date")),   safe_td(row.get("Actual End Time")),
        row.get("Time Variance"),
        safe_td(row.get("Duration")), safe_td(row.get("Total Duration")),
        row.get("Units"), row.get("Unit Type"), row.get("Work Type"),
        row.get("Shift Conditions"), row.get("Classification"),
        row.get("Number Of Breaks"), safe_td(row.get("Break Duration")),
        row.get("Consolidated With Timesheet Line Id"),
        row.get("Reviewed By"), row.get("Created Date UTC"),
    ]


def build_hours_excel(result: dict) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    shift_df = result["shift_df"]
    long_df  = result["long_shift_df"]
    bold_font   = Font(bold=True, name="Calibri", size=11)
    normal_font = Font(name="Calibri", size=11)
    red_font    = Font(bold=True, name="Calibri", size=11, color="FFFFFF")
    ws1 = wb.create_sheet("Long Shift")
    for ci, h in enumerate(HEADERS_29, 1):
        ws1.cell(1, ci, h if h else "").font = bold_font
    for ri, (_, row) in enumerate(long_df.iterrows() if not long_df.empty else iter([]), 2):
        for ci, v in enumerate(_raw_row_values(row), 1):
            ws1.cell(ri, ci, v if v is not None else "").font = normal_font
        ws1.cell(ri, 18).fill = HW_YELLOW
        if row.get("short_break_flag"):
            ws1.cell(ri, 12).fill = HW_YELLOW
    ws2 = wb.create_sheet("Weekly Hours")
    wh_df = shift_df.copy()
    wh_df.sort_values(["Employee Id", "shift_start"], inplace=True)
    wh_df.reset_index(drop=True, inplace=True)
    for ci, h in enumerate(HEADERS_29, 1):
        ws2.cell(1, ci, h if h else "").font = bold_font
    emp_totals = wh_df.groupby("Employee Id")["duration_hrs"].sum()
    exceed_ids = set(emp_totals[emp_totals > WEEKLY_HOURS_THRESHOLD].index.tolist())
    current_row = 2
    grand_total_secs = 0
    for emp_id, grp in wh_df.groupby("Employee Id", sort=False):
        exceeds    = emp_id in exceed_ids
        group_secs = int(grp["duration_hrs"].sum() * 3600)
        grand_total_secs += group_secs
        for _, row in grp.iterrows():
            for ci, v in enumerate(_raw_row_values(row), 1):
                ws2.cell(current_row, ci, v if v is not None else "").font = normal_font
            ws2.row_dimensions[current_row].outline_level = 2
            ws2.row_dimensions[current_row].hidden = not exceeds
            current_row += 1
        c_lbl = ws2.cell(current_row, 1, f"{emp_id} Total")
        c_dur = ws2.cell(current_row, 18, timedelta(seconds=group_secs))
        c_dur.number_format = "[h]:mm:ss"
        if exceeds:
            c_lbl.fill = HW_RED; c_dur.fill = HW_RED
            c_lbl.font = red_font; c_dur.font = red_font
        else:
            c_lbl.font = bold_font; c_dur.font = bold_font
        ws2.row_dimensions[current_row].outline_level = 1
        ws2.row_dimensions[current_row].hidden = False
        current_row += 1
    c_grand     = ws2.cell(current_row, 1, "Grand Total")
    c_grand_dur = ws2.cell(current_row, 18, timedelta(seconds=grand_total_secs))
    c_grand_dur.number_format = "[h]:mm:ss"
    if grand_total_secs / 3600 > WEEKLY_HOURS_THRESHOLD:
        c_grand.fill = HW_RED; c_grand_dur.fill = HW_RED
        c_grand.font = red_font; c_grand_dur.font = red_font
    else:
        c_grand.font = bold_font; c_grand_dur.font = bold_font
    ws2.row_dimensions[current_row].hidden = False
    ws2.sheet_properties.outlinePr.summaryBelow = True
    wb.create_sheet("Sheet1")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# KEYPAY LOCATION
# ═══════════════════════════════════════════════════════════════════════════════

KL_YELLOW = PatternFill('solid', fgColor='FFFF00')
KL_BOLD   = Font(bold=True)

KL_COL_WIDTHS = {
    'A': 11.82, 'B': 13.27, 'C': 14.45, 'D': 8.54,  'E': 12.54,
    'F': 10.27, 'G': 58.63, 'H': 10.45, 'I': 11.0,  'J': 10.0,
    'K': 11.0,  'L': 9.09,  'M': 10.55, 'N': 13.82, 'O': 5.45,
    'P': 9.45,  'Q': 27.09, 'R': 15.09, 'S': 12.54, 'T': 17.45,
    'U': 14.45, 'V': 33.82, 'W': 19.09, 'X': 16.55, 'Y': 18.27,
}


def kl_loc_prefix(location):
    if pd.isna(location) or not isinstance(location, str):
        return None
    parts = location.split('/')
    if len(parts) < 2:
        return None
    m = re.match(r'^([A-Z])\d{4,}', parts[1].strip(), re.IGNORECASE)
    return m.group(1).upper() if m else None


def kl_is_unallocated(loc):
    return kl_loc_prefix(loc) is None if not (pd.isna(loc) or not isinstance(loc, str)) else True


def kl_classify(row):
    status    = str(row.get('Status', '')).strip()
    location  = row.get('Location', '')
    work_type = str(row.get('Work Type', '')).strip() if pd.notna(row.get('Work Type')) else ''
    prefix    = kl_loc_prefix(location)
    if status == 'Processed':
        return 'exclude'
    rb = row.get('Reviewed By', '')
    if status == 'Approved' and not (pd.isna(rb) or str(rb).strip() == ''):
        fn = f"{str(row.get('First Name','')).strip()} {str(row.get('Surname','')).strip()}".strip()
        if str(rb).strip() == fn:
            return 'self_approved'
    if status == 'Approved' and work_type == 'Annual Leave Taken' and prefix == 'C':
        return 'al_c_costed'
    if status == 'Approved' and prefix == 'C':
        return 'exclude'
    if status == 'Approved' and prefix is None:
        return 'exclude'
    if status == 'Approved' and prefix != 'C' and work_type == '':
        return 'approved_non_c'
    if status == 'Approved' and prefix != 'C' and work_type != '':
        return 'exclude'
    if status == 'Submitted' and prefix is None:
        return 'unapproved_unallocated'
    if status == 'Submitted' and prefix is not None:
        return 'unapproved_allocated'
    return 'exclude'


def kl_build_excel(df_raw, results) -> bytes:
    wb = Workbook()
    headers = df_raw.columns.tolist()

    def write_section(ws, label, key):
        row_num = ws.max_row + 1
        cell = ws.cell(row=row_num, column=1, value=label)
        cell.fill = KL_YELLOW
        cell.font = KL_BOLD
        df_sec = results.get(key, pd.DataFrame())
        if df_sec.empty:
            ws.cell(row=ws.max_row + 1, column=1, value='N/A')
        else:
            for _, r in df_sec.iterrows():
                ws.append([None if pd.isna(r.get(h)) else r.get(h) for h in headers])

    ws_sum = wb.active
    ws_sum.title = 'Summary'
    ws_sum.append(headers)
    for ci in range(1, len(headers) + 1):
        ws_sum.cell(1, ci).font = KL_BOLD
    for col, w in KL_COL_WIDTHS.items():
        ws_sum.column_dimensions[col].width = w

    write_section(ws_sum, 'Approved – Unallocated',       'approved_unallocated')
    write_section(ws_sum, 'Approved – Non-C Locations',   'approved_non_c')
    write_section(ws_sum, 'Unapproved – Unallocated',     'unapproved_unallocated')
    write_section(ws_sum, 'Unapproved – Allocated',       'unapproved_allocated')
    write_section(ws_sum, 'Others - Self approved timesheets', 'self_approved')
    write_section(ws_sum, "Others - Approved AL but C costed (This is now updated to the employee's HOME location)", 'al_c_costed')

    ws_all = wb.create_sheet('All Timesheets')
    ws_all.append(headers)
    for ci in range(1, len(headers) + 1):
        ws_all.cell(1, ci).font = KL_BOLD
    for col, w in KL_COL_WIDTHS.items():
        ws_all.column_dimensions[col].width = w
    for _, r in df_raw.iterrows():
        ws_all.append([None if pd.isna(r.get(h)) else r.get(h) for h in headers])

    ws_det = wb.create_sheet('Report Details')
    ws_det.column_dimensions['A'].width = 20
    ws_det.column_dimensions['B'].width = 35
    now = datetime.now()
    now_str = now.strftime('%d/%m/%Y ') + now.strftime('%I:%M').lstrip('0') + '\u202f' + now.strftime('%p').lower()
    ws_det.append(['Report Details', None])
    ws_det.append(['Report Name', 'Timesheets report'])
    ws_det.append(['Date Generated', now_str])
    ws_det.append(['User', ''])
    ws_det.cell(1, 1).font = KL_BOLD

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def kl_classify_all(df_raw: pd.DataFrame) -> dict:
    df_kl = df_raw.copy()
    df_kl['_cat'] = df_kl.apply(kl_classify, axis=1)
    empty_df = pd.DataFrame(columns=df_kl.columns.drop('_cat'))

    def kl_sec(cat):
        return (df_kl[df_kl['_cat'] == cat].drop(columns='_cat')
                if cat in df_kl['_cat'].values else empty_df.copy())

    return {
        'approved_non_c':         kl_sec('approved_non_c'),
        'approved_unallocated':   kl_sec('approved_unallocated'),
        'unapproved_unallocated': kl_sec('unapproved_unallocated'),
        'unapproved_allocated':   kl_sec('unapproved_allocated'),
        'self_approved':          kl_sec('self_approved'),
        'al_c_costed':            kl_sec('al_c_costed'),
    }
