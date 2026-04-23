"""Chemika business logic.

Lifted from Handover_automations/Chemika/app.py as-is per the 'wrap, don't
rewrite' rule. Only the pure functions are copied; Streamlit calls are dropped.
"""

import io
import re

import openpyxl
import pandas as pd


EMPLOYEES = [
    {"row": 4,  "surname": "Ambrose",       "initial": "C", "category": "CASUAL",    "std_base": 165},
    {"row": 5,  "surname": "Bevan",          "initial": "H", "category": "PART-TIME", "std_base": 65},
    {"row": 6,  "surname": "Cantillon",      "initial": "J", "category": "",          "std_base": 162.5},
    {"row": 7,  "surname": "Charlton",       "initial": "H", "category": "",          "std_base": 162.5},
    {"row": 8,  "surname": "Cillian",        "initial": "A", "category": "",          "std_base": 162.5},
    {"row": 9,  "surname": "Darshika",       "initial": "",  "category": "SALARY",    "std_base": 162.5},
    {"row": 10, "surname": "Doherty",        "initial": "M", "category": "SALARY",    "std_base": 162.5},
    {"row": 11, "surname": "Doughty",        "initial": "S", "category": "CASUAL",    "std_base": 0},
    {"row": 12, "surname": "El Khoury",      "initial": "C", "category": "SALARY",    "std_base": 162.5},
    {"row": 13, "surname": "Emmanuel",       "initial": "M", "category": "SALARY",    "std_base": 162.5},
    {"row": 14, "surname": "Falzado",        "initial": "J", "category": "SALARY",    "std_base": 162.5},
    {"row": 15, "surname": "Horrigan",       "initial": "N", "category": "SALARY",    "std_base": 162.5},
    {"row": 16, "surname": "Jogia",          "initial": "V", "category": "SALARY",    "std_base": 162.5},
    {"row": 17, "surname": "Jones",          "initial": "A", "category": "SALARY",    "std_base": 162.5},
    {"row": 18, "surname": "Lauren",         "initial": "P", "category": "",          "std_base": 162.5},
    {"row": 19, "surname": "Lee",            "initial": "N", "category": "PART-TIME", "std_base": 108.3},
    {"row": 20, "surname": "LeStrange",      "initial": "S", "category": "SALARY",    "std_base": 162.5},
    {"row": 21, "surname": "LYAKHOVA",       "initial": "K", "category": "SALARY",    "std_base": 162.5},
    {"row": 22, "surname": "Manchanayake",   "initial": "T", "category": "PART-TIME", "std_base": 130},
    {"row": 23, "surname": "Mao",            "initial": "F", "category": "FULL-TIME", "std_base": 162.5},
    {"row": 24, "surname": "Micallef",       "initial": "C", "category": "SALARY",    "std_base": 162.5},
    {"row": 25, "surname": "Miller",         "initial": "R", "category": "SALARY",    "std_base": 162.5},
    {"row": 26, "surname": "Moeun",          "initial": "N", "category": "SALARY",    "std_base": 162.5},
    {"row": 27, "surname": "Parison",        "initial": "L", "category": "",          "std_base": 162.5},
    {"row": 28, "surname": "Piva",           "initial": "L", "category": "PART-TIME", "std_base": 130},
    {"row": 29, "surname": "Reardon",        "initial": "P", "category": "SALARY",    "std_base": 162.5},
    {"row": 30, "surname": "Rose",           "initial": "J", "category": "SALARY",    "std_base": 162.5},
    {"row": 31, "surname": "Sekuljica",      "initial": "D", "category": "FULL-TIME", "std_base": 162.5},
    {"row": 32, "surname": "Simpson",        "initial": "D", "category": "PART-TIME", "std_base": 97.5},
    {"row": 33, "surname": "Sor",            "initial": "S", "category": "CASUAL",    "std_base": 6},
    {"row": 34, "surname": "Watson",         "initial": "P", "category": "FULL-TIME", "std_base": 162.5},
]


def extract_numbers(s):
    if s is None:
        return 0
    result = "".join(c for c in str(s) if c.isdigit() or c == ".")
    return float(result) if result else 0


def safe_num(v):
    if v is None or str(v).strip().upper() == "N/A" or str(v).strip() == "":
        return 0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0


def extract_name_from_filename(filename):
    name_no_ext = filename.rsplit(".", 1)[0] if "." in filename else filename
    name_clean = name_no_ext.replace("_", " ").replace("-", "").strip()
    name_clean = re.sub(r"\s+", " ", name_clean)
    parts = name_clean.split()
    setstring = (parts[0] + " " + parts[1]) if len(parts) >= 2 else (parts[0] if parts else "")
    return setstring.strip()


def match_employee(setstring):
    setstring_clean = setstring.strip().lower()
    for emp in EMPLOYEES:
        surname = emp["surname"].strip()
        initial = emp["initial"].strip()
        if surname.lower() == setstring_clean:
            return emp
        if (surname + " " + initial).strip().lower() == setstring_clean:
            return emp
    return None


def parse_surname_initial(filename):
    name_no_ext = filename.rsplit(".", 1)[0] if "." in filename else filename
    name_clean = name_no_ext.replace("_", " ").replace("-", "").strip()
    name_clean = re.sub(r"\s+", " ", name_clean)
    parts = name_clean.split()
    month_names = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]
    name_parts = []
    for p in parts:
        if p.rstrip(".").lower() in month_names or p.isdigit():
            break
        name_parts.append(p)
    if len(name_parts) >= 2 and len(name_parts[-1]) == 1:
        return " ".join(name_parts[:-1]), name_parts[-1]
    elif name_parts:
        return " ".join(name_parts), ""
    return filename, ""


def process_timesheet(file_bytes, filename):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    setstring = extract_name_from_filename(filename)
    emp = match_employee(setstring)
    is_new = emp is None
    warnings, errors = [], []

    sheet_name = ws.title.replace(",", "").strip()
    file_parts = setstring.lower().split()
    sheet_parts = sheet_name.lower().split()
    if file_parts and sheet_parts and file_parts[0] != sheet_parts[0]:
        errors.append(f"Filename says **{setstring}** but sheet is named **{ws.title}** — possible wrong file or copy error")

    g1 = ws.cell(1, 7).value
    header_std = None
    if g1:
        nums = re.findall(r"[\d.]+", str(g1))
        if nums:
            try:
                header_std = float(nums[0])
            except ValueError:
                pass

    AL = PL = 0.0
    prev_month_row = 49
    for row in range(6, 56):
        a = ws.cell(row, 1).value
        if a and "previous" in str(a).lower():
            prev_month_row = row
            break

    daily_ot_total = total_daily_hours = 0.0
    dates_seen, unsigned_working_days, day_mismatch_rows, weekend_no_ot = {}, [], [], []

    scan_end = min(prev_month_row, 49)
    for row in range(6, scan_end):
        a     = ws.cell(row, 1).value
        b     = ws.cell(row, 2).value
        c_val = ws.cell(row, 3).value
        d     = ws.cell(row, 4).value
        e     = ws.cell(row, 5).value
        f_ot  = ws.cell(row, 6).value
        c = safe_num(c_val)

        if c > 0:
            total_daily_hours += c
        if f_ot is not None and isinstance(f_ot, (int, float)) and f_ot > 0:
            daily_ot_total += f_ot

        if b and hasattr(b, "strftime"):
            date_str = b.strftime("%Y-%m-%d")
            if date_str in dates_seen:
                warnings.append(f"**Duplicate date** {date_str} at rows {dates_seen[date_str]} and {row}")
            else:
                dates_seen[date_str] = row

            if a and str(a).strip().lower() in ("monday","tuesday","wednesday","thursday","friday","saturday","sunday"):
                if str(a).strip().lower() != b.strftime("%A").lower():
                    day_mismatch_rows.append(row)

            actual_day_name = b.strftime("%A")
            if actual_day_name in ("Saturday","Sunday") and c > 0:
                if f_ot is None or (isinstance(f_ot, (int, float)) and f_ot == 0):
                    weekend_no_ot.append(f"Row {row}: {actual_day_name} {date_str} — {c}hrs worked, no OT in column F")

        if a and str(a).strip().lower() in ("monday","tuesday","wednesday","thursday","friday"):
            if c > 0 and (d is None or str(d).strip() == "") and (e is None or str(e).strip() == ""):
                unsigned_working_days.append(row)

        if d is None:
            continue
        d_str = str(d).strip()
        d_lower = d_str.lower()

        if d_str in ("AL", "A/L") or d_lower == "annual leave":
            AL += c
        elif (d_str == "SL" or d_lower in ("carers leave", "sick")) and setstring != "LeStrange S" and len(d_str) < 11:
            PL += c
        elif "leave" in d_lower and d_lower not in ("annual leave",):
            PL += c

        if setstring == "LeStrange S" and d_str == "SL" and len(d_str) < 11:
            cell = ws.cell(row, 4)
            if cell.fill and cell.fill.start_color:
                rgb = str(cell.fill.start_color.rgb) if cell.fill.start_color.rgb else ""
                if "FFFF00" in rgb or "ffff00" in rgb.lower():
                    PL += c

        if "sl" in d_lower:
            num = extract_numbers(d_str)
            if num > 0 and d_str != "SL":
                PL += num

    if day_mismatch_rows:
        sample = ", ".join(str(r) for r in day_mismatch_rows[:5])
        extra = "..." if len(day_mismatch_rows) > 5 else ""
        warnings.append(f"**Day/date mismatch** on {len(day_mismatch_rows)} row(s) — day names don't match calendar dates (rows {sample}{extra})")
    for item in weekend_no_ot[:3]:
        warnings.append(f"**Weekend work without OT**: {item}")
    if unsigned_working_days:
        sample = ", ".join(str(r) for r in unsigned_working_days[:5])
        extra = "..." if len(unsigned_working_days) > 5 else ""
        warnings.append(f"**Missing sign-in** on {len(unsigned_working_days)} weekday(s) with hours recorded (rows {sample}{extra})")

    std_hrs = ot10 = ot15 = ot20 = 0
    for row in range(2, ws.max_row + 1):
        a = ws.cell(row, 1).value
        if a is None:
            continue
        a_str = str(a).strip()
        if a_str == "STD MONTHLY HOURS":
            std_hrs = safe_num(ws.cell(row, 3).value)
        elif a_str == "O/T X 1.0":
            ot10 = safe_num(ws.cell(row, 3).value)
        elif a_str == "O/T X 1.5":
            ot15 = safe_num(ws.cell(row, 3).value)
        elif a_str == "O/T X 2.0":
            ot20 = safe_num(ws.cell(row, 3).value)

    if header_std is not None and std_hrs > 0 and abs(header_std - std_hrs) > 0.01:
        errors.append(f"**STD hours mismatch**: Header says **{header_std}** but summary row says **{std_hrs}**")
    if std_hrs == 0:
        errors.append("**STD MONTHLY HOURS is 0 or missing** — cannot verify hours")

    summary_ot = ot10 + ot15 + ot20
    if (daily_ot_total > 0 or summary_ot > 0) and abs(daily_ot_total - summary_ot) > 0.5:
        warnings.append(f"**OT mismatch**: Daily OT entries total **{daily_ot_total:.1f}hrs** but summary OT rows total **{summary_ot:.1f}hrs** (diff: {abs(daily_ot_total - summary_ot):.1f}hrs)")

    if std_hrs > 0 and total_daily_hours > 0 and total_daily_hours < std_hrs * 0.5:
        warnings.append(f"**Low hours**: Total daily hours (**{total_daily_hours:.1f}**) is less than half of STD (**{std_hrs}**) — timesheet may be incomplete")

    lsl_hrs = 0
    e64 = ws.cell(64, 5).value
    if e64 and "lsl accrual" in str(e64).strip().lower():
        lsl_hrs = safe_num(str(ws.cell(64, 6).value).replace("hrs", ""))
    else:
        for check_row in [67, 68, 66, 57]:
            f_val = ws.cell(check_row, 6).value
            if f_val is not None and str(f_val).strip() and str(f_val).strip().upper() != "N/A":
                lsl_hrs = safe_num(str(f_val).replace("hrs", ""))
                break

    compass = safe_num(ws.cell(67, 3).value)
    wb.close()

    if is_new:
        surname, initial = parse_surname_initial(filename)
        emp = {"row": None, "surname": surname, "initial": initial, "category": "", "std_base": 0, "is_new": True}

    return {
        "filename": filename, "name": setstring, "matched": True, "is_new": is_new,
        "employee": emp, "std_hrs": std_hrs, "lsl_hrs": lsl_hrs,
        "ot10": ot10, "ot15": ot15, "ot20": ot20,
        "al": AL, "pl": PL, "compass": compass,
        "warnings": warnings, "errors": errors,
    }


def build_payroll_output(results, month_label) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Main"

    ws.cell(1, 2, f"PAYROLL TIMESHEET - Month of {month_label}")
    ws.cell(1, 2).font = openpyxl.styles.Font(bold=True, size=14)

    headers = {
        1: "No", 2: "EMPLOYEE", 3: "", 4: "Start date", 5: "No. Yrs",
        6: "Category", 7: "Days p/w", 8: "STD HRS", 9: "LSL Hours",
        10: "Normal Hours", 11: "Car Allow", 12: "First Aid Allow",
        13: "O/T 1.0", 14: "O/T 1.5", 15: "O/T 2.0",
        16: "Travel @85c per km", 17: "Annual Leave", 18: "Personal/Sick Leave",
        19: "Compass Leave", 20: "LSL Leave", 21: "LWOP", 22: "Bonus"
    }
    hf = openpyxl.styles.Font(bold=True, size=10)
    hfill = openpyxl.styles.PatternFill("solid", fgColor="D9E1F2")
    for col, hdr in headers.items():
        cell = ws.cell(3, col, hdr)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", wrap_text=True)

    result_map, new_results = {}, []
    for r in results:
        if r["matched"]:
            if r.get("is_new"):
                new_results.append(r)
            else:
                result_map[r["employee"]["surname"].strip().lower()] = r

    for emp in EMPLOYEES:
        row = emp["row"]
        ws.cell(row, 1, row - 3)
        ws.cell(row, 2, emp["surname"])
        ws.cell(row, 3, emp["initial"])
        ws.cell(row, 6, emp["category"] if emp["category"] else None)
        key = emp["surname"].strip().lower()
        if key in result_map:
            r = result_map[key]
            ws.cell(row, 8,  r["std_hrs"])
            ws.cell(row, 9,  r["lsl_hrs"])
            ws.cell(row, 13, r["ot10"] if r["ot10"] != 0 else "N/A")
            ws.cell(row, 14, r["ot15"] if r["ot15"] != 0 else "N/A")
            ws.cell(row, 15, r["ot20"] if r["ot20"] != 0 else "N/A")
            ws.cell(row, 17, r["al"])
            ws.cell(row, 18, r["pl"])
            ws.cell(row, 19, r["compass"])

    last_row = max(e["row"] for e in EMPLOYEES)
    new_fill = openpyxl.styles.PatternFill("solid", fgColor="FFF2CC")
    for idx, r in enumerate(sorted(new_results, key=lambda x: x["employee"]["surname"])):
        row = last_row + 1 + idx
        emp_r = r["employee"]
        ws.cell(row, 1, row - 3)
        ws.cell(row, 2, emp_r["surname"])
        ws.cell(row, 3, emp_r["initial"])
        ws.cell(row, 8,  r["std_hrs"])
        ws.cell(row, 9,  r["lsl_hrs"])
        ws.cell(row, 13, r["ot10"] if r["ot10"] != 0 else "N/A")
        ws.cell(row, 14, r["ot15"] if r["ot15"] != 0 else "N/A")
        ws.cell(row, 15, r["ot20"] if r["ot20"] != 0 else "N/A")
        ws.cell(row, 17, r["al"])
        ws.cell(row, 18, r["pl"])
        ws.cell(row, 19, r["compass"])
        for col in range(1, 23):
            ws.cell(row, col).fill = new_fill

    total_rows = last_row + len(new_results)
    col_widths = {1:5,2:18,3:4,4:12,5:8,6:12,7:8,8:10,9:10,10:12,11:10,12:12,13:8,14:8,15:8,16:16,17:12,18:16,19:12,20:10,21:8,22:8}
    for col, w in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    for row in range(4, total_rows + 1):
        for col in [8, 9, 13, 14, 15, 17, 18, 19, 20]:
            cell = ws.cell(row, col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    ws_issues = wb.create_sheet("Data Issues")
    ih_font = openpyxl.styles.Font(bold=True, size=11, color="FFFFFF")
    ih_fill = openpyxl.styles.PatternFill("solid", fgColor="4472C4")
    for col_idx, hdr in enumerate(["Employee", "Severity", "Issue"], 1):
        cell = ws_issues.cell(1, col_idx, hdr)
        cell.font = ih_font
        cell.fill = ih_fill

    err_fill   = openpyxl.styles.PatternFill("solid", fgColor="FFC7CE")
    err_font   = openpyxl.styles.Font(color="9C0006")
    warn_fill  = openpyxl.styles.PatternFill("solid", fgColor="FFEB9C")
    warn_font  = openpyxl.styles.Font(color="9C6500")
    issue_row  = 2
    all_r = list(result_map.values()) + new_results
    for r in sorted(all_r, key=lambda x: x["employee"]["surname"]):
        emp_name = f"{r['employee']['surname']} {r['employee']['initial']}".strip()
        for err in r.get("errors", []):
            clean = re.sub(r"\*\*", "", err)
            ws_issues.cell(issue_row, 1, emp_name)
            ws_issues.cell(issue_row, 2, "ERROR")
            ws_issues.cell(issue_row, 3, clean)
            for c in range(1, 4):
                ws_issues.cell(issue_row, c).fill = err_fill
                ws_issues.cell(issue_row, c).font = err_font
            issue_row += 1
        for warn in r.get("warnings", []):
            clean = re.sub(r"\*\*", "", warn)
            ws_issues.cell(issue_row, 1, emp_name)
            ws_issues.cell(issue_row, 2, "WARNING")
            ws_issues.cell(issue_row, 3, clean)
            for c in range(1, 4):
                ws_issues.cell(issue_row, c).fill = warn_fill
                ws_issues.cell(issue_row, c).font = warn_font
            issue_row += 1

    if issue_row == 2:
        ws_issues.cell(2, 1, "No issues found")
        ws_issues.cell(2, 1).font = openpyxl.styles.Font(color="006100")

    ws_issues.column_dimensions["A"].width = 20
    ws_issues.column_dimensions["B"].width = 12
    ws_issues.column_dimensions["C"].width = 90

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def txt_format_date(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, str):
        return val.strip()
    if isinstance(val, pd.Timestamp):
        return f"{val.month}/{val.day}/{val.year}"
    try:
        dt = pd.to_datetime(val)
        return f"{dt.month}/{dt.day}/{dt.year}"
    except Exception:
        return str(val)


def txt_clean_num(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    f = float(val)
    return str(int(f)) if f == int(f) else str(f)


def build_txt(
    df: pd.DataFrame,
    memo: str,
    due_date: int,
    due_days: int,
    tax_code: str,
    account: str,
) -> bytes:
    TAB, CRLF = "\t", "\r\n"
    required = ["Date", "Sub Total", "GST", "Company Name", "Invoice Number"]
    for col in required:
        if col not in df.columns:
            raise ValueError(f"Missing required column: '{col}'")
    df = df.copy()
    if "Other" not in df.columns:
        df["Other"] = ""
    df = df.sort_values(
        by=["Company Name", "Invoice Number"],
        key=lambda col: col.str.lower() if col.name == "Company Name"
                        else pd.to_numeric(col, errors="coerce").fillna(0),
    )
    header = TAB.join([
        "Date", "Sub Total", "Other", "GST", "Company Name", "Invoice Number",
        "Memo", "TT Ex GST", "TT Inc GST", "Due Date", "Due Days", "Tax Code", "Account",
    ])
    blank = TAB * 12
    lines = [header]
    for _, row in df.iterrows():
        date_str = txt_format_date(row["Date"])
        sub_total = txt_clean_num(row["Sub Total"])
        other = txt_clean_num(row.get("Other", ""))
        gst = txt_clean_num(row["GST"])
        company = str(row["Company Name"]).strip()
        invoice = str(int(row["Invoice Number"])) if pd.notna(row["Invoice Number"]) else ""
        try:
            tt_ex_str = txt_clean_num(float(row["Sub Total"]))
            tt_inc_str = txt_clean_num(float(row["Sub Total"]) + float(row["GST"]))
        except Exception:
            tt_ex_str, tt_inc_str = sub_total, ""
        lines.append(TAB.join([
            date_str, sub_total, other, gst, company, invoice,
            memo, tt_ex_str, tt_inc_str,
            str(due_date), str(due_days), tax_code, account,
        ]))
        lines.append(blank)
    return (CRLF.join(lines)).encode("utf-8")
